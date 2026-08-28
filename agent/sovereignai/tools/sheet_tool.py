"""
sheet_tool.py — openpyxl-based spreadsheet read/write tools.

Preserves existing formatting and formulas on read-modify-write cycles.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

from sovereignai.tools.base import Tool, ToolResult
from sovereignai.tools.fs_tools import _validate_path


class SheetRead(Tool):
    name = "sheet_read"
    description = "Read cell values and formulas from an .xlsx spreadsheet."
    categories = ["spreadsheet", "planning", "general"]
    json_schema = {
        "type": "object",
        "properties": {
            "path":       {"type": "string",  "description": "Path to .xlsx file"},
            "sheet_name": {"type": "string",  "description": "Sheet name (first sheet if omitted)"},
            "max_rows":   {"type": "integer", "description": "Max rows to read (default: 500)"},
        },
        "required": ["path"],
    }

    def run(self, path: str, sheet_name: str | None = None, max_rows: int = 500) -> ToolResult:
        p = _validate_path(path)
        if p is None:
            return ToolResult.fail(f"Path '{path}' outside workspace.")
        if not p.exists():
            return ToolResult.fail(f"File not found: {path}")

        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(p), data_only=False)
            ws = wb[sheet_name] if sheet_name else wb.active

            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=False)):
                if i >= max_rows:
                    break
                row_data = []
                for cell in row:
                    row_data.append({
                        "cell": cell.coordinate,
                        "value": cell.value,
                        "formula": cell.value if str(cell.value or "").startswith("=") else None,
                    })
                rows.append(row_data)

            return ToolResult.ok({
                "sheet": ws.title,
                "rows": rows,
                "max_row": ws.max_row,
                "max_col": ws.max_column,
                "sheets": wb.sheetnames,
            })
        except ImportError:
            return ToolResult.fail("openpyxl not installed. Run: pip install openpyxl")
        except Exception as e:
            return ToolResult.fail(str(e))


class SheetWrite(Tool):
    name = "sheet_write"
    description = "Update specific cells in an existing .xlsx spreadsheet."
    categories = ["spreadsheet", "planning"]
    json_schema = {
        "type": "object",
        "properties": {
            "path":       {"type": "string", "description": "Path to .xlsx file"},
            "sheet_name": {"type": "string", "description": "Sheet name (first sheet if omitted)"},
            "updates": {
                "type": "array",
                "description": "List of {cell, value} or {cell, formula} updates",
                "items": {
                    "type": "object",
                    "properties": {
                        "cell":    {"type": "string", "description": "Cell ref e.g. 'A1'"},
                        "value":   {"description": "New value"},
                        "formula": {"type": "string", "description": "Formula string e.g. '=SUM(A1:A10)'"},
                    },
                    "required": ["cell"],
                },
            },
        },
        "required": ["path", "updates"],
    }

    def run(self, path: str, updates: list[dict], sheet_name: str | None = None) -> ToolResult:
        p = _validate_path(path)
        if p is None:
            return ToolResult.fail(f"Path '{path}' outside workspace.")
        if not p.exists():
            return ToolResult.fail(f"File not found: {path}")

        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(p))
            ws = wb[sheet_name] if sheet_name else wb.active

            applied = []
            for u in updates:
                cell_ref = u.get("cell")
                formula = u.get("formula")
                value = u.get("value")
                if formula:
                    ws[cell_ref] = formula
                else:
                    ws[cell_ref] = value
                applied.append(cell_ref)

            wb.save(str(p))
            return ToolResult.ok({"updated_cells": applied, "path": str(p)}, file_path=str(p))
        except ImportError:
            return ToolResult.fail("openpyxl not installed.")
        except Exception as e:
            return ToolResult.fail(str(e))


class SheetCreate(Tool):
    name = "sheet_create"
    description = "Create a new .xlsx spreadsheet with headers and rows."
    categories = ["spreadsheet", "planning", "general"]
    json_schema = {
        "type": "object",
        "properties": {
            "path":       {"type": "string",  "description": "Output file path (.xlsx)"},
            "sheet_name": {"type": "string",  "description": "Sheet name (default: Sheet1)"},
            "headers":    {"type": "array",   "items": {"type": "string"}, "description": "Column headers"},
            "rows": {
                "type": "array",
                "description": "List of row arrays (each row is a list of values)",
                "items": {"type": "array"},
            },
        },
        "required": ["path", "headers"],
    }

    def run(self, path: str, headers: list[str], rows: list[list] | None = None, sheet_name: str = "Sheet1") -> ToolResult:
        p = _validate_path(path)
        if p is None:
            return ToolResult.fail(f"Path '{path}' outside workspace.")

        try:
            import openpyxl
            from openpyxl.styles import Font
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name

            # Write headers in bold
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)

            # Write data rows
            for row_idx, row in enumerate(rows or [], 2):
                for col_idx, val in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=val)

            p.parent.mkdir(parents=True, exist_ok=True)
            wb.save(str(p))
            return ToolResult.ok({
                "path": str(p),
                "sheets": [sheet_name],
                "rows": len(rows or []),
                "cols": len(headers),
            }, file_path=str(p))
        except ImportError:
            return ToolResult.fail("openpyxl not installed.")
        except Exception as e:
            return ToolResult.fail(str(e))

